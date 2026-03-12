"""
Data Quality Engine  ─  MetaHarvest v3
Compares source schema against a target (Collibra or another scan),
runs structural, completeness, volume, and integrity checks.
"""
from __future__ import annotations
from typing import List, Dict, Optional, Callable
from datetime import datetime

from app.models import (
    TableMeta, ColumnMeta, ScanResult,
    DQTableResult, DQColumnResult, DQRunResult
)
from app.config import DQ_THRESHOLDS


# ══════════════════════════════════════════════════════════════
#  DQ Engine
# ══════════════════════════════════════════════════════════════
class DQEngine:
    def __init__(self):
        self._cancel = False

    def cancel(self): self._cancel = True

    def run(
        self,
        source_scan:   ScanResult,
        target_scan:   Optional[ScanResult]  = None,
        collibra_meta: Optional[Dict]         = None,
        progress_cb:   Optional[Callable]     = None,
        checks:        Optional[List[str]]    = None,
    ) -> DQRunResult:
        """
        source_scan  : The scan you just ran (from any connector)
        target_scan  : Optional second scan to compare against (e.g. same tables in different env)
        collibra_meta: Optional dict of {table_name: [col_names]} pulled from Collibra
        checks       : List of check IDs to run (None = all)
        """
        result = DQRunResult(
            scan_id    = source_scan.scan_id,
            source_env = source_scan.environment,
            target_env = target_scan.environment if target_scan else "collibra",
        )
        total = len(source_scan.tables)

        for i, src_table in enumerate(source_scan.tables):
            if self._cancel: break
            if not src_table.selected: continue
            if progress_cb:
                progress_cb(f"DQ check: {src_table.table_name}…", i, total)

            # Find matching target table
            tgt_table = None
            if target_scan:
                tgt_table = self._find_matching(src_table, target_scan.tables)

            # Build column-level comparison
            tbl_result = self._check_table(src_table, tgt_table, collibra_meta, checks)
            result.tables.append(tbl_result)

            if tbl_result.status == "ok":    result.summary_pass += 1
            elif tbl_result.status == "warn": result.summary_warn += 1
            else:                             result.summary_fail += 1

        result.finished_at = datetime.now().isoformat()
        result.status      = (
            "fail" if result.summary_fail > 0
            else "warn" if result.summary_warn > 0
            else "ok"
        )
        if progress_cb:
            progress_cb(f"DQ complete — {result.summary_pass} pass, "
                        f"{result.summary_warn} warn, {result.summary_fail} fail",
                        total, total)
        return result

    # ── Table check ──────────────────────────────────────────
    def _check_table(
        self,
        src:           TableMeta,
        tgt:           Optional[TableMeta],
        collibra_meta: Optional[Dict],
        checks:        Optional[List[str]],
    ) -> DQTableResult:
        all_checks = checks is None
        res = DQTableResult(
            table_name       = src.table_name,
            source_path      = src.full_path,
            target_path      = tgt.full_path if tgt else "",
            source_row_count = src.row_count,
            target_row_count = tgt.row_count if tgt else None,
            source_col_count = src.col_count,
            target_col_count = tgt.col_count if tgt else 0,
        )

        src_cols = {c.name.lower(): c for c in src.columns}
        tgt_cols = {c.name.lower(): c for c in tgt.columns} if tgt else {}

        # ── Collibra drift check ──────────────────────────────
        collibra_cols: Dict[str, str] = {}
        if collibra_meta:
            key = src.table_name.lower()
            if key in collibra_meta:
                collibra_cols = {c.lower(): "" for c in collibra_meta[key]}

        # ── Determine universe of columns ─────────────────────
        all_col_names = set(src_cols) | set(tgt_cols) | set(collibra_cols)

        for col_name in sorted(all_col_names):
            src_col = src_cols.get(col_name)
            tgt_col = tgt_cols.get(col_name)

            cr = DQColumnResult(
                column_name = col_name,
                in_source   = col_name in src_cols,
                in_target   = col_name in tgt_cols if tgt else True,
            )

            if src_col:
                cr.source_type     = src_col.data_type
                cr.source_nullable = src_col.nullable
            if tgt_col:
                cr.target_type     = tgt_col.data_type
                cr.target_nullable = tgt_col.nullable

            # ── Structural checks ─────────────────────────────
            if (all_checks or "orphan_cols" in (checks or [])):
                if not cr.in_source:
                    cr.issues.append(f"Column present in target but MISSING from source")
            if (all_checks or "new_cols" in (checks or [])):
                if cr.in_source and not cr.in_target and tgt:
                    cr.issues.append(f"NEW column in source — not in target")

            # ── Type compatibility ────────────────────────────
            if (all_checks or "type_compat" in (checks or [])):
                if src_col and tgt_col:
                    cr.type_compatible = self._types_compatible(
                        src_col.data_type, tgt_col.data_type)
                    if not cr.type_compatible:
                        cr.issues.append(
                            f"Type mismatch: source={src_col.data_type} "
                            f"target={tgt_col.data_type}")

            # ── Collibra drift ────────────────────────────────
            if collibra_cols and cr.in_source and col_name not in collibra_cols:
                cr.issues.append("Column exists in source but NOT in Collibra catalog")

            res.columns.append(cr)

        # ── Volume check ──────────────────────────────────────
        if (all_checks or "row_count" in (checks or [])):
            if src.row_count and tgt and tgt.row_count:
                src_r = src.row_count
                tgt_r = tgt.row_count
                if src_r > 0:
                    variance = abs(src_r - tgt_r) / src_r
                    res.row_variance_pct = round(variance * 100, 2)
                    if variance > DQ_THRESHOLDS["row_count_fail"]:
                        res.checks_failed += 1
                    elif variance > DQ_THRESHOLDS["row_count_warn"]:
                        res.checks_warned += 1
                    else:
                        res.checks_passed += 1

        # ── Field count check ─────────────────────────────────
        if (all_checks or "field_count" in (checks or [])):
            if tgt and src.col_count != tgt.col_count:
                diff = abs(src.col_count - tgt.col_count)
                if diff > 5:
                    res.checks_failed += 1
                elif diff > 0:
                    res.checks_warned += 1
            else:
                res.checks_passed += 1

        # ── Tally column-level issues ─────────────────────────
        for cr in res.columns:
            if not cr.in_source:
                res.checks_failed += 1
            elif not cr.in_target and tgt:
                res.checks_warned += 1
            elif cr.issues:
                if any("mismatch" in i.lower() or "missing" in i.lower()
                       for i in cr.issues):
                    res.checks_warned += 1
            else:
                res.checks_passed += 1

        # ── Overall table status ──────────────────────────────
        if res.checks_failed > 0:
            res.status = "fail"
        elif res.checks_warned > 0:
            res.status = "warn"
        else:
            res.status = "ok"

        return res

    def _find_matching(self, src: TableMeta,
                       candidates: List[TableMeta]) -> Optional[TableMeta]:
        """Find best matching table in target scan."""
        # Exact name match first
        for t in candidates:
            if t.table_name.lower() == src.table_name.lower():
                return t
        # Schema + name
        for t in candidates:
            if (t.schema.lower()     == src.schema.lower() and
                t.table_name.lower() == src.table_name.lower()):
                return t
        return None

    @staticmethod
    def _types_compatible(src_type: str, tgt_type: str) -> bool:
        """Broad compatibility — numeric stays numeric, string stays string, etc."""
        def category(dt: str) -> str:
            dt = dt.upper()
            if any(x in dt for x in ("INT","BIGINT","SMALLINT","TINYINT","NUMBER","NUMERIC","DECIMAL","FLOAT","DOUBLE","REAL","MONEY")):
                return "numeric"
            if any(x in dt for x in ("VARCHAR","CHAR","TEXT","STRING","CLOB","NVARCHAR","NCHAR")):
                return "string"
            if any(x in dt for x in ("DATE","TIME","TIMESTAMP","DATETIME")):
                return "datetime"
            if any(x in dt for x in ("BOOL","BIT","BOOLEAN")):
                return "boolean"
            if any(x in dt for x in ("BLOB","BINARY","VARBINARY","BYTEA","RAW")):
                return "binary"
            return "other"
        return category(src_type) == category(tgt_type)


# ── DQ Report Export ───────────────────────────────────────────
def export_dq_report(result: DQRunResult, path: str):
    """Export DQ results to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        wb = openpyxl.Workbook()

        # ── Summary sheet ──────────────────────────────────────
        ws = wb.active
        ws.title = "DQ Summary"
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 30

        header_fill  = PatternFill("solid", fgColor="0D1320")
        ok_fill      = PatternFill("solid", fgColor="0D2E1F")
        warn_fill    = PatternFill("solid", fgColor="2A1800")
        fail_fill    = PatternFill("solid", fgColor="2E0D14")
        header_font  = Font(bold=True, color="00C2FF")

        ws.append(["Table", "Status", "Src Cols", "Tgt Cols",
                   "Row Var %", "Issues"])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for t in result.tables:
            issues = (f"{len(t.dropped_cols)} dropped cols"
                      if t.dropped_cols else "")
            if t.type_mismatches:
                issues += (", " if issues else "") + f"{len(t.type_mismatches)} type mismatches"
            row = [
                t.table_name, t.status.upper(),
                t.source_col_count, t.target_col_count,
                f"{t.row_variance_pct:.1f}%" if t.row_variance_pct is not None else "N/A",
                issues or "OK",
            ]
            ws.append(row)
            fill = (ok_fill if t.status == "ok"
                    else warn_fill if t.status == "warn" else fail_fill)
            for cell in ws[ws.max_row]:
                cell.fill = fill

        # ── Detail sheet ───────────────────────────────────────
        ws2 = wb.create_sheet("Column Details")
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 30
        ws2.column_dimensions["C"].width = 20
        ws2.column_dimensions["D"].width = 20
        ws2.column_dimensions["E"].width = 15
        ws2.column_dimensions["F"].width = 50
        ws2.append(["Table", "Column", "Source Type", "Target Type",
                    "Status", "Issues"])
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font

        for t in result.tables:
            for c in t.columns:
                ws2.append([
                    t.table_name, c.column_name,
                    c.source_type, c.target_type,
                    c.status.upper(),
                    " | ".join(c.issues) if c.issues else "OK",
                ])
                fill = (ok_fill if c.status == "ok"
                        else warn_fill if c.status in ("warn","new")
                        else fail_fill)
                for cell in ws2[ws2.max_row]:
                    cell.fill = fill

        wb.save(path)
        return True, path
    except Exception as e:
        return False, str(e)
